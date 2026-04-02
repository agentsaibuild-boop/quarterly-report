#!/usr/bin/env python3
"""
Събира данни за тримесечния отчет:
- Чете разходите/приходите директно от Тримесечен отчет.xlsx
- Сканира всички проекти (git commits, активност)
Резултатът се записва в output/data.json за анализ от Claude Code.
"""

import json
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Инсталирам openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                    "--break-system-packages", "--quiet"], check=True)
    import openpyxl

BASE_DIR = Path(__file__).parent
XLSX = BASE_DIR / "Тримесечен отчет.xlsx"
CONFIG = BASE_DIR / "config.json"
OUTPUT = BASE_DIR / "output" / "data.json"


# ── Excel ─────────────────────────────────────────────────────────────────────

def read_excel():
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)

    # Лист "2026" — приходи и разходи по месеци
    ws = wb["2026"]
    rows = list(ws.iter_rows(values_only=True))

    # Ред 4 (index 3) съдържа датите на колоните D,E,F
    header_row = rows[3]
    months = []
    for cell in header_row[3:6]:
        if isinstance(cell, datetime):
            months.append(cell.strftime("%Y-%m"))
        elif cell:
            months.append(str(cell))

    income = {}
    expenses = {}
    for row in rows[4:]:
        label = row[1]
        if not label or not isinstance(label, str):
            continue
        label = label.strip()
        if not label:
            continue

        values = {}
        for i, m in enumerate(months):
            v = row[3 + i]
            if isinstance(v, (int, float)):
                values[m] = round(v, 2)

        if "ПРИХОДИ" in label.upper() or label.startswith("Приход"):
            income[label] = values
        elif label not in ("РАЗХОДИ", "Общо разходи", "КОНТРОЛИРУЕМ РЕЗУЛТАТ"):
            if any(v != 0 for v in values.values()):
                expenses[label] = values

    # Лист "AI инструменти"
    ws2 = wb["AI инструменти"]
    ai_tools = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            ai_tools.append({
                "tool": row[0],
                "plan": row[1] or "",
                "price_per_month": row[2] if isinstance(row[2], (int, float)) else 0,
                "purpose": row[3] or "",
            })

    return {
        "months": months,
        "income": income,
        "expenses": expenses,
        "ai_tools": ai_tools,
        "totals_per_month": _totals(expenses, months),
    }


def _totals(expenses, months):
    totals = {m: 0.0 for m in months}
    for values in expenses.values():
        for m, v in values.items():
            totals[m] = round(totals.get(m, 0) + v, 2)
    return totals


# ── Git проекти ───────────────────────────────────────────────────────────────

def git_stats(repo_path, since, until):
    path = Path(repo_path)
    if not (path / ".git").exists():
        return {"error": "не е git репо"}

    def run(cmd):
        try:
            r = subprocess.run(cmd, cwd=repo_path, capture_output=True,
                               text=True, timeout=15)
            return r.stdout.strip()
        except Exception:
            return ""

    af = [f"--since={since}", f"--until={until}"]

    log = run(["git", "log", "--oneline"] + af)
    lines_log = [l for l in log.splitlines() if l.strip()]
    commits = len(lines_log)
    messages = [l.split(" ", 1)[1] if " " in l else l for l in lines_log[:15]]

    stat = run(["git", "log", "--numstat", "--pretty="] + af)
    added = removed = 0
    for line in stat.splitlines():
        p = line.split("\t")
        if len(p) == 3:
            try:
                added += int(p[0])
                removed += int(p[1])
            except ValueError:
                pass

    last = run(["git", "log", "-1", "--format=%ci %s"])
    branch = run(["git", "rev-parse", "--abbrev-ref", "HEAD"])

    return {
        "commits": commits,
        "lines_added": added,
        "lines_removed": removed,
        "last_commit": last,
        "branch": branch,
        "recent_commits": messages,
    }


def folder_stats(folder_path):
    path = Path(folder_path)
    if not path.exists():
        return {"error": "не съществува"}
    files = total_size = newest = 0
    for f in path.rglob("*"):
        if f.is_file() and ".git" not in f.parts:
            files += 1
            try:
                s = f.stat()
                total_size += s.st_size
                if s.st_mtime > newest:
                    newest = s.st_mtime
            except OSError:
                pass
    return {
        "files": files,
        "size_kb": round(total_size / 1024, 1),
        "last_modified": datetime.fromtimestamp(newest).strftime("%Y-%m-%d") if newest else "?",
    }


# ── GitHub авто-откриване ────────────────────────────────────────────────────

REPOS_DIR = Path.home() / "projects"  # папка за всички клонирани репота


def get_github_repos() -> list[dict]:
    """Взема всички репота на потребителя от GitHub чрез gh CLI."""
    result = subprocess.run(
        ["gh", "repo", "list", "--json", "name,sshUrl,description,isPrivate",
         "--limit", "200"],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print("  ВНИМАНИЕ: gh не е логнат. Пусни: gh auth login")
        return []
    return json.loads(result.stdout)


def sync_all_repos(clone_dir: Path) -> list[dict]:
    """Клонира/обновява всички GitHub репота и връща списък за сканиране."""
    clone_dir.mkdir(parents=True, exist_ok=True)
    repos = get_github_repos()
    if not repos:
        return []

    projects = []
    for r in repos:
        name = r["name"]
        path = clone_dir / name

        if not (path / ".git").exists():
            print(f"  Клонирам {name}...")
            subprocess.run(["git", "clone", "--quiet", r["sshUrl"], str(path)],
                           capture_output=True)
        else:
            print(f"  pull {name}...")
            subprocess.run(["git", "pull", "--quiet"],
                           cwd=str(path), capture_output=True)

        projects.append({
            "name": name,
            "description": r.get("description") or "",
            "type": "git",
            "path": str(path),
        })

    return projects


# ── Main ──────────────────────────────────────────────────────────────────────

def collect(months: int = 3):
    config = json.loads(CONFIG.read_text(encoding="utf-8"))
    end = datetime.now()
    start = end - timedelta(days=months * 30)
    since = start.strftime("%Y-%m-%d")
    until = end.strftime("%Y-%m-%d")

    print(f"Период: {since} → {until}")

    # Авто-открива и обновява всички GitHub репота
    clone_dir = Path(config.get("clone_dir", str(REPOS_DIR)))
    print(f"Синхронизирам GitHub репота → {clone_dir}")
    github_projects = sync_all_repos(clone_dir)

    # Добавя ръчно конфигурираните (папки без git, локални проекти извън GitHub)
    manual = [p for p in config.get("manual_projects", [])]

    all_projects = github_projects + manual

    # Финанси от Excel
    print("Чета Excel...")
    finances = read_excel()

    # Сканира всички проекти
    print(f"\nСканирам {len(all_projects)} проекта...")
    projects = []
    for p in all_projects:
        print(f"  {p['name']}...")
        if p.get("type", "git") == "git":
            stats = git_stats(p["path"], since, until)
        else:
            stats = folder_stats(p["path"])
        projects.append({
            "name": p["name"],
            "description": p.get("description", ""),
            "type": p.get("type", "git"),
            "stats": stats,
        })

    data = {
        "generated_at": datetime.now().isoformat(),
        "period": {"from": since, "to": until, "months": months},
        "finances": finances,
        "projects": projects,
    }

    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nДанни записани: {OUTPUT}")
    print("Готово. Покажи data.json на Claude Code за анализ.")
    return data


if __name__ == "__main__":
    m = int(sys.argv[1]) if len(sys.argv) > 1 else 3
    collect(m)
